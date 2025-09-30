import os
import sys
import sqlite3
import requests
import json
import time
import subprocess
from pathlib import Path
from typing import Dict, List, Any, Tuple
import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, PieChart, LineChart

class ValidationResults:    
    def __init__(self):
        self.results = {}
        self.total_score = 0
        self.max_score = 0
    
    def add_test(self, test_name: str, passed: bool, score: int, max_score: int, details: str = ""):
        self.results[test_name] = {
            'passed': passed,
            'score': score if passed else 0,
            'max_score': max_score,
            'details': details
        }
        self.total_score += (score if passed else 0)
        self.max_score += max_score
    
    def print_summary(self):
        print("\n" + "="*60)
        print("🏦 BANK AI DATA ANALYST - VALIDATION REPORT")
        print("="*60)
        
        for test_name, result in self.results.items():
            status = "✅ PASS" if result['passed'] else "❌ FAIL"
            score = f"{result['score']}/{result['max_score']}"
            print(f"{status} {test_name:<40} {score:>8}")
            if result['details']:
                print(f"    📝 {result['details']}")
        
        print("-"*60)
        percentage = (self.total_score / self.max_score * 100) if self.max_score > 0 else 0
        print(f"JAMI BALL: {self.total_score}/{self.max_score} ({percentage:.1f}%)")
        
        if percentage >= 90:
            print("🏆 EXCELLENT - Barcha talablar bajarilgan!")
        elif percentage >= 75:
            print("✅ GOOD - Ko'pchilik talablar bajarilgan")
        elif percentage >= 50:
            print("⚠️  AVERAGE - Ba'zi talablar bajarilmagan")
        else:
            print("❌ NEEDS WORK - Ko'p talablar bajarilmagan")
        
        print("="*60)

class TZValidator:    
    def __init__(self):
        self.results = ValidationResults()
        self.db_path = "bank_data.db"
        self.web_url = "http://localhost:5000"
        self.reports_dir = "reports"
    
    def validate_all(self) -> ValidationResults:
        print("🏦 Bank AI Data Analyst - TZ Validation")
        print("Topshiriq talablarini tekshirilmoqda...")
        print()
        
        self._validate_database()
        self._validate_llm_integration()
        self._validate_sql_generation()
        self._validate_excel_export()
        self._validate_interface()      
        self._validate_bonus_features()
        
        return self.results
    
    def _validate_database(self):
        print("1️⃣ Ma'lumotlar bazasi tekshirilmoqda...")
        
        if not os.path.exists(self.db_path):
            self.results.add_test("Database file exists", False, 0, 5, 
                                f"Database fayli topilmadi: {self.db_path}")
            return
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in cursor.fetchall()]
            
            required_tables = ['clients', 'accounts', 'transactions']
            missing_tables = [t for t in required_tables if t not in tables]
            
            if missing_tables:
                self.results.add_test("Required tables exist", False, 0, 5,
                                    f"Jadvallar topilmadi: {missing_tables}")
            else:
                self.results.add_test("Required tables exist", True, 5, 5)
            
            total_records = 0
            
            cursor.execute("SELECT COUNT(*) FROM clients")
            clients_count = cursor.fetchone()[0]
            total_records += clients_count
            
            cursor.execute("SELECT COUNT(*) FROM accounts")
            accounts_count = cursor.fetchone()[0]
            total_records += accounts_count
            
            cursor.execute("SELECT COUNT(*) FROM transactions")
            transactions_count = cursor.fetchone()[0]
            total_records += transactions_count
            
            if total_records >= 1000000:
                self.results.add_test("1M+ records", True, 10, 10,
                                    f"Jami {total_records:,} yozuv")
            elif total_records >= 500000:
                self.results.add_test("1M+ records", True, 7, 10,
                                    f"Jami {total_records:,} yozuv (1M dan kam)")
            else:
                self.results.add_test("1M+ records", False, 0, 10,
                                    f"Jami {total_records:,} yozuv (juda kam)")
            
            schema_valid = True
            
            cursor.execute("PRAGMA table_info(clients)")
            client_columns = [row[1] for row in cursor.fetchall()]
            required_client_cols = ['id', 'name', 'birth_date', 'region']
            if not all(col in client_columns for col in required_client_cols):
                schema_valid = False
            
            cursor.execute("PRAGMA table_info(accounts)")
            account_columns = [row[1] for row in cursor.fetchall()]
            required_account_cols = ['id', 'client_id', 'balance', 'open_date']
            if not all(col in account_columns for col in required_account_cols):
                schema_valid = False
            
            cursor.execute("PRAGMA table_info(transactions)")
            transaction_columns = [row[1] for row in cursor.fetchall()]
            required_transaction_cols = ['id', 'account_id', 'amount', 'date', 'type']
            if not all(col in transaction_columns for col in required_transaction_cols):
                schema_valid = False
            
            self.results.add_test("Database schema valid", schema_valid, 5 if schema_valid else 0, 5)
            
            cursor.execute("SELECT COUNT(*) FROM clients WHERE region IN ('Toshkent', 'Samarqand', 'Buxoro')")
            regional_data = cursor.fetchone()[0]
            
            if regional_data > 0:
                self.results.add_test("Regional data exists", True, 5, 5,
                                    f"{regional_data} ta viloyat ma'lumotlari")
            else:
                self.results.add_test("Regional data exists", False, 0, 5,
                                    "Viloyat ma'lumotlari topilmadi")
            
            conn.close()
            
        except Exception as e:
            self.results.add_test("Database access", False, 0, 25,
                                f"Database xatosi: {str(e)}")
    
    def _validate_llm_integration(self):
        print("2️⃣ LLM integratsiyasi tekshirilmoqda...")
        
        try:
            response = requests.get("http://localhost:11434/api/tags", timeout=5)
            if response.status_code == 200:
                models = response.json()
                model_names = [model['name'] for model in models.get('models', [])]
                
                if any('llama' in model.lower() for model in model_names):
                    self.results.add_test("Ollama with Llama model", True, 10, 10,
                                        f"Models: {', '.join(model_names)}")
                else:
                    self.results.add_test("Ollama with Llama model", True, 5, 10,
                                        f"Ollama ishlaydi, lekin Llama model yo'q")
            else:
                self.results.add_test("Ollama service", False, 0, 10,
                                    f"Ollama API xatosi: {response.status_code}")
        
        except requests.exceptions.RequestException:
            self.results.add_test("Ollama service", False, 0, 10,
                                "Ollama service ishlamaydi yoki mavjud emas")
        
        try:
            from bank_analyst import LLMQueryGenerator
            llm_gen = LLMQueryGenerator()
            
            test_prompt = "Toshkent viloyatidagi mijozlar sonini ko'rsat"
            sql_query = llm_gen.generate_sql(test_prompt)
            
            if sql_query and "SELECT" in sql_query.upper():
                self.results.add_test("LLM query generation", True, 10, 10,
                                    f"SQL yaratildi: {sql_query[:50]}...")
            else:
                self.results.add_test("LLM query generation", False, 0, 10,
                                    "SQL query yaratilmadi")
        
        except Exception as e:
            self.results.add_test("LLM integration code", False, 0, 10,
                                f"LLM kodi xatosi: {str(e)}")
    
    def _validate_sql_generation(self):
        print("3️⃣ SQL query generation tekshirilmoqda...")
        
        try:
            from bank_analyst import BankAnalystAssistant
            
            assistant = BankAnalystAssistant()
            assistant.db_manager.connect()
            
            test_cases = [
                {
                    'prompt': "Toshkent viloyatidagi mijozlar sonini ko'rsat",
                    'expected_keywords': ['SELECT', 'COUNT', 'clients', 'region', 'Toshkent']
                },
                {
                    'prompt': "2024 yildagi jami tranzaksiyalar summasini ko'rsat",
                    'expected_keywords': ['SELECT', 'SUM', 'transactions', '2024']
                },
                {
                    'prompt': "Eng ko'p balansga ega 10 ta hisobni ko'rsat",
                    'expected_keywords': ['SELECT', 'balance', 'accounts', 'ORDER BY', 'LIMIT', '10']
                }
            ]
            
            passed_tests = 0
            
            for i, test_case in enumerate(test_cases):
                try:
                    result = assistant.process_query(test_case['prompt'])
                    
                    if result['success']:
                        sql_query = result['sql_query'].upper()
                        keywords_found = sum(1 for keyword in test_case['expected_keywords'] 
                                           if keyword.upper() in sql_query)
                        
                        if keywords_found >= len(test_case['expected_keywords']) * 0.7:  # 70% keywords
                            passed_tests += 1
                            print(f"  ✅ Test {i+1}: {test_case['prompt'][:30]}...")
                        else:
                            print(f"  ❌ Test {i+1}: Keywords kam: {keywords_found}/{len(test_case['expected_keywords'])}")
                    else:
                        print(f"  ❌ Test {i+1}: Query xatosi: {result.get('error', 'Unknown')}")
                
                except Exception as e:
                    print(f"  ❌ Test {i+1}: Exception: {str(e)}")
            
            if passed_tests == len(test_cases):
                self.results.add_test("SQL generation accuracy", True, 20, 20,
                                    f"Barcha {len(test_cases)} test muvaffaqiyatli")
            elif passed_tests >= len(test_cases) * 0.7:
                self.results.add_test("SQL generation accuracy", True, 15, 20,
                                    f"{passed_tests}/{len(test_cases)} test muvaffaqiyatli")
            else:
                self.results.add_test("SQL generation accuracy", False, 0, 20,
                                    f"Faqat {passed_tests}/{len(test_cases)} test muvaffaqiyatli")
            
            assistant.close()
            
        except Exception as e:
            self.results.add_test("SQL generation functionality", False, 0, 20,
                                f"SQL generation xatosi: {str(e)}")
    
    def _validate_excel_export(self):
        print("4️⃣ Excel export va grafiklar tekshirilmoqda...")
        
        try:
            from bank_analyst import BankAnalystAssistant
            
            assistant = BankAnalystAssistant()
            assistant.db_manager.connect()
            
            test_query = "Har bir viloyatdagi mijozlar sonini ko'rsat"
            filepath = assistant.generate_report(test_query, 'bar')
            
            if filepath and os.path.exists(filepath):
                self.results.add_test("Excel file generation", True, 5, 5,
                                    f"Fayl yaratildi: {os.path.basename(filepath)}")
                
                try:
                    workbook = openpyxl.load_workbook(filepath)
                    worksheet = workbook.active
                    
                    if worksheet.max_row > 1 and worksheet.max_column > 1:
                        self.results.add_test("Excel data content", True, 5, 5,
                                            f"{worksheet.max_row-1} qator ma'lumot")
                    else:
                        self.results.add_test("Excel data content", False, 0, 5,
                                            "Excel faylda ma'lumot yo'q")
                    
                    charts = list(worksheet._charts)
                    if charts:
                        chart_types = [type(chart).__name__ for chart in charts]
                        self.results.add_test("Excel charts", True, 5, 5,
                                            f"Grafiklar: {', '.join(chart_types)}")
                    else:
                        self.results.add_test("Excel charts", False, 0, 5,
                                            "Excel faylda grafiklar yo'q")
                    
                    workbook.close()
                    
                except Exception as e:
                    self.results.add_test("Excel content validation", False, 0, 5,
                                        f"Excel tahlil xatosi: {str(e)}")
                
            else:
                self.results.add_test("Excel file generation", False, 0, 15,
                                    "Excel fayl yaratilmadi")
            
            assistant.close()
            
        except Exception as e:
            self.results.add_test("Excel export functionality", False, 0, 15,
                                f"Excel export xatosi: {str(e)}")
    
    def _validate_interface(self):
        print("5️⃣ Interfeys tekshirilmoqda...")
        
        if os.path.exists("bank_analyst.py"):
            try:
                result = subprocess.run([
                    sys.executable, "bank_analyst.py", "--query", 
                    "SELECT COUNT(*) as total FROM clients"
                ], capture_output=True, text=True, timeout=30)
                
                if result.returncode == 0:
                    self.results.add_test("CLI interface", True, 5, 5,
                                        "CLI query muvaffaqiyatli")
                else:
                    self.results.add_test("CLI interface", False, 0, 5,
                                        f"CLI xatosi: {result.stderr}")
                
            except Exception as e:
                self.results.add_test("CLI interface", False, 0, 5,
                                    f"CLI test xatosi: {str(e)}")
        else:
            self.results.add_test("CLI interface", False, 0, 5,
                                "bank_analyst.py topilmadi")
        
        try:
            response = requests.get(f"{self.web_url}/", timeout=10)
            if response.status_code == 200:
                self.results.add_test("Web interface", True, 5, 5,
                                    "Web UI ishlaydi")
                
                try:
                    stats_response = requests.get(f"{self.web_url}/api/stats", timeout=5)
                    if stats_response.status_code == 200:
                        stats_data = stats_response.json()
                        if 'clients' in stats_data and 'transactions' in stats_data:
                            print(f"  ✅ Stats API: {stats_data['clients']} mijoz, {stats_data['transactions']} tranzaksiya")
                        else:
                            print("  ⚠️  Stats API: Ma'lumot to'liq emas")
                    
                    examples_response = requests.get(f"{self.web_url}/api/examples", timeout=5)
                    if examples_response.status_code == 200:
                        examples_data = examples_response.json()
                        if 'examples' in examples_data and len(examples_data['examples']) > 0:
                            print(f"  ✅ Examples API: {len(examples_data['examples'])} misol")
                        else:
                            print("  ⚠️  Examples API: Misol yo'q")
                    
                except Exception as e:
                    print(f"  ⚠️  API test xatosi: {str(e)}")
                    
            else:
                self.results.add_test("Web interface", False, 0, 5,
                                    f"Web UI xatosi: {response.status_code}")
        
        except requests.exceptions.RequestException as e:
            self.results.add_test("Web interface", False, 0, 5,
                                "Web UI ishlamaydi yoki mavjud emas")
    
    def _validate_bonus_features(self):
        print("6️⃣ Bonus features tekshirilmoqda...")
        
        if os.path.exists("Dockerfile"):
            self.results.add_test("Dockerfile exists", True, 3, 3,
                                "Docker support mavjud")
            
            try:
                result = subprocess.run([
                    "docker", "build", "-t", "bank-analyst-test", "."
                ], capture_output=True, text=True, timeout=300)
                
                if result.returncode == 0:
                    self.results.add_test("Docker build", True, 2, 2,
                                        "Docker image build muvaffaqiyatli")
                    
                    subprocess.run(["docker", "rmi", "bank-analyst-test"], 
                                 capture_output=True)
                else:
                    self.results.add_test("Docker build", False, 0, 2,
                                        "Docker build xatosi")
                    
            except Exception as e:
                self.results.add_test("Docker build", False, 0, 2,
                                    f"Docker test xatosi: {str(e)}")
        else:
            self.results.add_test("Dockerfile exists", False, 0, 5,
                                "Docker support yo'q")
        
        if os.path.exists("web_app.py"):
            self.results.add_test("Web UI implementation", True, 3, 3,
                                "Web interfeys yaratilgan")
        else:
            self.results.add_test("Web UI implementation", False, 0, 3,
                                "Web interfeys yo'q")
        
        if os.path.exists("production_config.py"):
            self.results.add_test("Production configuration", True, 2, 2,
                                "Production sozlamalar mavjud")
        else:
            self.results.add_test("Production configuration", False, 0, 2,
                                "Production config yo'q")
    
    def _check_code_quality(self):
        print("📝 Kod sifati tekshirilmoqda...")
        
        required_files = [
            "bank_analyst.py",
            "requirements.txt", 
            "README.md"
        ]
        
        missing_files = [f for f in required_files if not os.path.exists(f)]
        
        if not missing_files:
            self.results.add_test("Required files", True, 5, 5,
                                "Barcha kerakli fayllar mavjud")
        else:
            self.results.add_test("Required files", False, 0, 5,
                                f"Fayllar topilmadi: {missing_files}")
        
        if os.path.exists("README.md"):
            with open("README.md", 'r', encoding='utf-8') as f:
                readme_content = f.read()
                
            readme_sections = [
                "o'rnatish", "ishlatish", "docker", "demo"
            ]
            
            sections_found = sum(1 for section in readme_sections 
                               if section.lower() in readme_content.lower())
            
            if sections_found >= 3:
                self.results.add_test("README quality", True, 3, 3,
                                    f"{sections_found}/4 bo'lim mavjud")
            else:
                self.results.add_test("README quality", False, 0, 3,
                                    f"README to'liq emas: {sections_found}/4")
        
        try:
            from bank_analyst import BankAnalystAssistant, DatabaseManager, LLMQueryGenerator, ExcelExporter
            self.results.add_test("Code structure", True, 2, 2,
                                "Asosiy klasslar mavjud")
        except ImportError as e:
            self.results.add_test("Code structure", False, 0, 2,
                                f"Import xatosi: {str(e)}")

def run_full_validation():
    """To'liq validation test"""
    print("🏦 BANK AI DATA ANALYST - FINAL VALIDATION")
    print("=" * 60)
    print("Test Assignment talablarini tekshirish...")
    print()
    
    validator = TZValidator()
    results = validator.validate_all()
    validator._check_code_quality()
    results.print_summary()
    
    print("\n📋 TAVSIYALAR:")
    print("-" * 40)
    
    failed_tests = [name for name, result in results.results.items() 
                   if not result['passed']]
    
    if not failed_tests:
        print("🎉 Barcha testlar muvaffaqiyatli o'tdi!")
        print("✅ Loyiha production uchun tayyor!")
    else:
        print("⚠️  Quyidagi masalalarni hal qiling:")
        for test_name in failed_tests:
            details = results.results[test_name]['details']
            print(f"  • {test_name}: {details}")
    
    print(f"\n🏆 Yakuniy ball: {results.total_score}/{results.max_score}")
    
    return results

def create_demo_report():
    """Demo uchun test hisobot yaratish"""
    print("\n📊 Demo hisobot yaratilmoqda...")
    
    try:
        from bank_analyst import BankAnalystAssistant
        
        assistant = BankAnalystAssistant()
        assistant.db_manager.connect()
        
        demo_queries = [
            ("Viloyatlar bo'yicha mijozlar", "Har bir viloyatdagi mijozlar sonini ko'rsat", "pie"),
            ("Eng katta balanslar", "Eng ko'p balansga ega 10 ta hisobni ko'rsat", "bar"),
            ("Hisob turlari", "Har bir hisob turida qancha hisob borligini ko'rsat", "pie")
        ]
        
        created_reports = []
        
        for title, query, chart_type in demo_queries:
            try:
                print(f"  📈 {title}...")
                filepath = assistant.generate_report(query, chart_type)
                if filepath:
                    created_reports.append((title, filepath))
                    print(f"    ✅ Yaratildi: {os.path.basename(filepath)}")
                else:
                    print(f"    ❌ Yaratilmadi")
            except Exception as e:
                print(f"    ❌ Xato: {str(e)}")
        
        assistant.close()
        
        if created_reports:
            print(f"\n🎉 {len(created_reports)} ta demo hisobot yaratildi:")
            for title, filepath in created_reports:
                print(f"  📄 {title}: {filepath}")
        else:
            print("\n❌ Demo hisobotlar yaratilmadi")
        
        return created_reports
        
    except Exception as e:
        print(f"❌ Demo hisobot yaratishda xato: {str(e)}")
        return []

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Bank AI Data Analyst - TZ Validation')
    parser.add_argument('--quick', action='store_true', help='Tez test (Docker build o\'tkazmaslik)')
    parser.add_argument('--demo', action='store_true', help='Demo hisobotlar yaratish')
    parser.add_argument('--web-check', action='store_true', help='Faqat web interfeys tekshirish')
    
    args = parser.parse_args()
    
    if args.web_check:
        print("🌐 Web interfeys tekshirilmoqda...")
        validator = TZValidator()
        validator._validate_interface()
        validator.results.print_summary()
        return
    
    if args.demo:
        create_demo_report()
        return
    
    results = run_full_validation()
    
    if args.demo:
        create_demo_report()
    
    percentage = (results.total_score / results.max_score * 100) if results.max_score > 0 else 0
    if percentage >= 75:
        sys.exit(0)  
    else:
        sys.exit(1) 

if __name__ == "__main__":
    main()